import openpyxl
import os


# Função para exibir o menu
def exibir_menu():
    menu = """
    [d] Depositar
    [s] Sacar
    [e] Extrato
    [c] Cadastrar Usuário
    [a] Cadastrar Conta Bancária
    [u] Exibir Usuários e Contas
    [x] Exportar Dados para Excel
    [q] Sair
    => """
    return input(menu).lower()


# Função para carregar os dados dos usuários e contas do arquivo Excel
def carregar_dados(filename="dados_usuarios.xlsx"):
    if not os.path.exists(filename):
        return {}, {}

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    usuarios = {}
    contas = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        nome, cpf, endereco, numero_conta, saldo, extrato, numero_saques = row

        if cpf not in usuarios:
            usuarios[cpf] = {"nome": nome, "data_nascimento": "", "endereco": endereco, "contas": []}

        conta = {
            "numero": numero_conta,
            "saldo": saldo,
            "extrato": extrato if extrato else "",
            "numero_saques": numero_saques if numero_saques else 0
        }
        usuarios[cpf]["contas"].append(conta)
        contas[numero_conta] = conta

    return usuarios, contas


# Função para salvar os dados dos usuários e contas no arquivo Excel
def salvar_dados(filename="dados_usuarios.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Usuários e Contas"

    # Cabeçalhos
    sheet.append(["Nome", "CPF", "Endereço", "Número da Conta", "Saldo", "Extrato", "Número de Saques"])

    for cpf, dados_usuario in usuarios.items():
        for conta in dados_usuario["contas"]:
            sheet.append([dados_usuario["nome"], cpf, dados_usuario["endereco"], conta["numero"], conta["saldo"],
                          conta["extrato"], conta["numero_saques"]])

    workbook.save(filename)
    print(f"Dados salvos com sucesso em {filename}")


# Função para depósito
def depositar(conta):
    valor = float(input("Informe o valor do depósito: "))

    if valor > 0:
        conta["saldo"] += valor
        conta["extrato"] += f"Depósito: R$ {valor:.2f}\n"
        salvar_dados()
    else:
        print("Operação falhou! O valor informado é inválido.")

    return conta


# Função para saque
def sacar(conta, limite, LIMITE_SAQUES):
    valor = float(input("Informe o valor do saque: "))

    excedeu_saldo = valor > conta["saldo"]
    excedeu_limite = valor > limite
    excedeu_saques = conta["numero_saques"] >= LIMITE_SAQUES

    if valor <= 0:
        print("Operação falhou! O valor informado é inválido.")
    elif excedeu_saldo:
        print("Operação falhou! Você não tem saldo suficiente.")
    elif excedeu_limite:
        print("Operação falhou! O valor do saque excede o limite.")
    elif excedeu_saques:
        print("Operação falhou! Número máximo de saques excedido.")
    else:
        conta["saldo"] -= valor
        conta["extrato"] += f"Saque: R$ {valor:.2f}\n"
        conta["numero_saques"] += 1
        salvar_dados()

    return conta


# Função para exibir o extrato
def exibir_extrato(conta):
    print("\n================ EXTRATO ================")
    if conta["extrato"]:
        print(conta["extrato"])
    else:
        print("Não foram realizadas movimentações.")
    print(f"\nSaldo: R$ {conta['saldo']:.2f}")
    print("==========================================")


# Função para cadastrar usuário
def cadastrar_usuario(usuarios):
    cpf = input("Informe o CPF (somente números): ")
    if cpf in usuarios:
        print("Usuário já cadastrado!")
        return usuarios

    nome = input("Informe o nome completo: ")
    data_nascimento = input("Informe a data de nascimento (dd/mm/aaaa): ")
    endereco = input("Informe o endereço (logradouro, número - bairro - cidade/UF): ")

    usuarios[cpf] = {"nome": nome, "data_nascimento": data_nascimento, "endereco": endereco, "contas": []}
    salvar_dados()
    print("Usuário cadastrado com sucesso!")

    return usuarios


# Função para cadastrar conta bancária
def cadastrar_conta(contas, usuarios):
    cpf = input("Informe o CPF do usuário: ")
    if cpf not in usuarios:
        print("Usuário não encontrado!")
        return contas

    numero_conta = len(contas) + 1
    conta = {"numero": numero_conta, "saldo": 0, "extrato": "", "numero_saques": 0}
    contas[numero_conta] = conta
    usuarios[cpf]["contas"].append(conta)
    salvar_dados()
    print(f"Conta {numero_conta} cadastrada com sucesso para o usuário {usuarios[cpf]['nome']}!")

    return contas


# Função para selecionar uma conta
def selecionar_conta(usuarios):
    cpf = input("Informe o CPF do usuário: ")
    if cpf not in usuarios:
        print("Usuário não encontrado!")
        return None, None

    print("Contas do usuário:")
    for conta in usuarios[cpf]["contas"]:
        print(f"Conta {conta['numero']}: Saldo R$ {conta['saldo']:.2f}")

    numero_conta = int(input("Informe o número da conta: "))
    for conta in usuarios[cpf]["contas"]:
        if conta["numero"] == numero_conta:
            return cpf, conta

    print("Conta não encontrada!")
    return None, None


# Função para exibir todos os usuários e suas contas
def exibir_usuarios_e_contas(usuarios):
    print("\n================ USUÁRIOS E CONTAS ================")
    for cpf, dados_usuario in usuarios.items():
        print(f"Usuário: {dados_usuario['nome']} - CPF: {cpf}")
        print("Contas:")
        for conta in dados_usuario["contas"]:
            print(f"  Conta {conta['numero']}: Saldo R$ {conta['saldo']:.2f}")
    print("===================================================")


# Função para exportar dados dos usuários para uma planilha Excel
def exportar_dados_para_excel(usuarios, filename="dados_usuarios_export.xlsx"):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Usuários e Contas"

        # Cabeçalhos
        sheet.append(["Nome", "CPF", "Endereço", "Número da Conta", "Saldo"])

        for cpf, dados_usuario in usuarios.items():
            for conta in dados_usuario["contas"]:
                sheet.append([dados_usuario["nome"], cpf, dados_usuario["endereco"], conta["numero"], conta["saldo"]])

        workbook.save(filename)
        print(f"Dados exportados com sucesso para {filename}")
    except PermissionError:
        print(f"Erro: Permissão negada ao tentar salvar {filename}. Tentando salvar com um nome diferente.")
        alternative_filename = "dados_usuarios_backup.xlsx"
        try:
            workbook.save(alternative_filename)
            print(f"Dados exportados com sucesso para {alternative_filename}")
        except Exception as e:
            print(f"Erro ao exportar dados para Excel: {e}")
    except Exception as e:
        print(f"Erro ao exportar dados para Excel: {e}")


# Variáveis globais
usuarios, contas = carregar_dados()

limite = 1500
LIMITE_SAQUES = 10

# Loop principal
while True:
    try:
        opcao = exibir_menu()

        if opcao == "d":
            cpf, conta = selecionar_conta(usuarios)
            if conta:
                conta = depositar(conta)

        elif opcao == "s":
            cpf, conta = selecionar_conta(usuarios)
            if conta:
                conta = sacar(conta, limite, LIMITE_SAQUES)

        elif opcao == "e":
            cpf, conta = selecionar_conta(usuarios)
            if conta:
                exibir_extrato(conta)

        elif opcao == "c":
            usuarios = cadastrar_usuario(usuarios)

        elif opcao == "a":
            contas = cadastrar_conta(contas, usuarios)

        elif opcao == "u":
            exibir_usuarios_e_contas(usuarios)

        elif opcao == "x":
            exportar_dados_para_excel(usuarios)

        elif opcao == "q":
            salvar_dados()
            break

        else:
            print("Operação inválida, por favor selecione novamente a operação desejada.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
        salvar_dados()
