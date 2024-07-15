import openpyxl
import os


class Usuario:
    def __init__(self, nome, cpf, data_nascimento, endereco):
        self.nome = nome
        self.cpf = cpf
        self.data_nascimento = data_nascimento
        self.endereco = endereco
        self.contas = []

    def adicionar_conta(self, conta):
        self.contas.append(conta)


class Conta:
    def __init__(self, numero, saldo=0, extrato="", numero_saques=0):
        self.numero = numero
        self.saldo = saldo
        self.extrato = extrato
        self.numero_saques = numero_saques

    def depositar(self, valor):
        if valor > 0:
            self.saldo += valor
            self.extrato += f"Depósito: R$ {valor:.2f}\n"
        else:
            print("Operação falhou! O valor informado é inválido.")

    def sacar(self, valor, limite, LIMITE_SAQUES):
        excedeu_saldo = valor > self.saldo
        excedeu_limite = valor > limite
        excedeu_saques = self.numero_saques >= LIMITE_SAQUES

        if valor <= 0:
            print("Operação falhou! O valor informado é inválido.")
        elif excedeu_saldo:
            print("Operação falhou! Você não tem saldo suficiente.")
        elif excedeu_limite:
            print("Operação falhou! O valor do saque excede o limite.")
        elif excedeu_saques:
            print("Operação falhou! Número máximo de saques excedido.")
        else:
            self.saldo -= valor
            self.extrato += f"Saque: R$ {valor:.2f}\n"
            self.numero_saques += 1

    def exibir_extrato(self):
        print("\n================ EXTRATO ================")
        if self.extrato:
            print(self.extrato)
        else:
            print("Não foram realizadas movimentações.")
        print(f"\nSaldo: R$ {self.saldo:.2f}")
        print("==========================================")


class Banco:
    def __init__(self, filename="dados_usuarios.xlsx"):
        self.filename = filename
        self.usuarios, self.contas = self.carregar_dados()

    def carregar_dados(self):
        if not os.path.exists(self.filename):
            return {}, {}

        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook.active

        usuarios = {}
        contas = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome, cpf, endereco, numero_conta, saldo, extrato, numero_saques = row

            if cpf not in usuarios:
                usuarios[cpf] = Usuario(nome, cpf, "", endereco)

            conta = Conta(numero_conta, saldo, extrato if extrato else "", numero_saques if numero_saques else 0)
            usuarios[cpf].adicionar_conta(conta)
            contas[numero_conta] = conta

        return usuarios, contas

    def salvar_dados(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Usuários e Contas"

            # Cabeçalhos
            sheet.append(["Nome", "CPF", "Endereço", "Número da Conta", "Saldo", "Extrato", "Número de Saques"])

            for usuario in self.usuarios.values():
                for conta in usuario.contas:
                    sheet.append([usuario.nome, usuario.cpf, usuario.endereco, conta.numero, conta.saldo,
                                  conta.extrato, conta.numero_saques])

            workbook.save(self.filename)
            print(f"Dados salvos com sucesso em {self.filename}")
        except Exception as e:
            print(f"Erro ao salvar dados: {e}")

    def cadastrar_usuario(self):
        try:
            cpf = input("Informe o CPF (somente números): ")
            if cpf in self.usuarios:
                print("Usuário já cadastrado!")
                return

            nome = input("Informe o nome completo: ")
            data_nascimento = input("Informe a data de nascimento (dd/mm/aaaa): ")
            endereco = input("Informe o endereço (logradouro, número - bairro - cidade/UF): ")

            self.usuarios[cpf] = Usuario(nome, cpf, data_nascimento, endereco)
            print("Usuário criado, tentando salvar dados...")
            self.salvar_dados()
            print("Usuário cadastrado com sucesso!")
        except Exception as e:
            print(f"Erro ao cadastrar usuário: {e}")

    def cadastrar_conta(self):
        try:
            cpf = input("Informe o CPF do usuário: ")
            if cpf not in self.usuarios:
                print("Usuário não encontrado!")
                return

            numero_conta = len(self.contas) + 1
            conta = Conta(numero_conta)
            self.contas[numero_conta] = conta
            self.usuarios[cpf].adicionar_conta(conta)
            print("Conta criada, tentando salvar dados...")
            self.salvar_dados()
            print(f"Conta {numero_conta} cadastrada com sucesso para o usuário {self.usuarios[cpf].nome}!")
        except Exception as e:
            print(f"Erro ao cadastrar conta: {e}")

    def selecionar_conta(self):
        cpf = input("Informe o CPF do usuário: ")
        if cpf not in self.usuarios:
            print("Usuário não encontrado!")
            return None, None

        print("Contas do usuário:")
        for conta in self.usuarios[cpf].contas:
            print(f"Conta {conta.numero}: Saldo R$ {conta.saldo:.2f}")

        numero_conta = int(input("Informe o número da conta: "))
        for conta in self.usuarios[cpf].contas:
            if conta.numero == numero_conta:
                return cpf, conta

        print("Conta não encontrada!")
        return None, None

    def exibir_usuarios_e_contas(self):
        print("\n================ USUÁRIOS E CONTAS ================")
        for usuario in self.usuarios.values():
            print(f"Usuário: {usuario.nome} - CPF: {usuario.cpf}")
            print("Contas:")
            for conta in usuario.contas:
                print(f"  Conta {conta.numero}: Saldo R$ {conta.saldo:.2f}")
        print("===================================================")

    def exportar_dados_para_excel(self, filename="dados_usuarios_export.xlsx"):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Usuários e Contas"

            # Cabeçalhos
            sheet.append(["Nome", "CPF", "Endereço", "Número da Conta", "Saldo"])

            for usuario in self.usuarios.values():
                for conta in usuario.contas:
                    sheet.append([usuario.nome, usuario.cpf, usuario.endereco, conta.numero, conta.saldo])

            workbook.save(filename)
            print(f"Dados exportados com sucesso para {filename}")
        except PermissionError:
            print(f"Erro: Permissão negada ao tentar salvar {filename}. Tentando salvar com um nome diferente.")
            alternative_filename = "dados_usuarios_backup.xlsx"
            try:
                workbook.save(alternative_filename)
                print(f"Dados exportados com sucesso para {alternative_filename}")
            except Exception as e:
                print(f"Erro ao exportar dados para Excel: {e}")
        except Exception as e:
            print(f"Erro ao exportar dados para Excel: {e}")


def exibir_menu():
    menu = """
    [d] Depositar
    [s] Sacar
    [e] Extrato
    [c] Cadastrar Usuário
    [a] Cadastrar Conta Bancária
    [u] Exibir Usuários e Contas
    [x] Exportar Dados para Excel
    [q] Sair
    => """
    return input(menu).lower()


def main():
    banco = Banco()

    limite = 1500
    LIMITE_SAQUES = 10

    while True:
        try:
            opcao = exibir_menu()

            if opcao == "d":
                cpf, conta = banco.selecionar_conta()
                if conta:
                    valor = float(input("Informe o valor do depósito: "))
                    conta.depositar(valor)
                    banco.salvar_dados()

            elif opcao == "s":
                cpf, conta = banco.selecionar_conta()
                if conta:
                    valor = float(input("Informe o valor do saque: "))
                    conta.sacar(valor, limite, LIMITE_SAQUES)
                    banco.salvar_dados()

            elif opcao == "e":
                cpf, conta = banco.selecionar_conta()
                if conta:
                    conta.exibir_extrato()

            elif opcao == "c":
                banco.cadastrar_usuario()

            elif opcao == "a":
                banco.cadastrar_conta()

            elif opcao == "u":
                banco.exibir_usuarios_e_contas()

            elif opcao == "x":
                banco.exportar_dados_para_excel()

            elif opcao == "q":
                banco.salvar_dados()
                break

            else:
                print("Operação inválida, por favor selecione novamente a operação desejada.")
        except Exception as e:
            print(f"Ocorreu um erro: {e}")
            banco.salvar_dados()


if __name__ == "__main__":
    main()
